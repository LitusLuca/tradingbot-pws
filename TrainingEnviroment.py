import math
from datetime import datetime, timedelta

import mysql.connector
import numpy


from _env import password, user
from openpyxl import Workbook, load_workbook
from os.path import exists


class StockSimulation:
    def __init__(self, instrument, timeSpan, episodeStart, results):
        self.time = 0
        self.windowSize = 10
        self.inventory = []
        self.invested = 0.0
        self.profit = 0.0
        self.instrument = instrument #TODO
        self.actionSpace = 3
        self.features = 3
        #self.inputSpace = self.windowSize * 5 + 3 #+3: invested profit and inventorie lenght
        self.inputSpace = [self.windowSize, self.features]
        self.trainingdata = self._getData(instrument, episodeStart-timedelta(days=self.windowSize), timeSpan)
        _, self.instrumentValue, _ = self.trainingdata[self.time]
        print(self.instrumentValue)

    #values for storing results
        if exists(results):
            self.wb = load_workbook(results)
        else:
            self.wb = Workbook()
        self.ws = self.wb.active
        self.specialparameter = 0.0
        self.specialparameterIterator = 0
        self.episode = 1
        self.ws["A1"] = "episode \\ parameter"
        self.ws["B1"] = self.specialparameter
        self.wb.save(results)
        self.resultsfile = results
        self.epsilon = 0
        self.startDate = episodeStart


    def _getData(self, instrument, startDate, length):
        cnx = mysql.connector.connect(user=user,password=password, database='indexes')
        cursor = cnx.cursor()

        table = instrument

        dataquery = ("SELECT i.Date, i.close, tnx.close FROM `{}` AS i, tnx "
            "WHERE i.date BETWEEN %s AND %s "
            "AND tnx.date = i.date").format(table)

        first_date = startDate
        last_date = first_date + timedelta(days=length)

        cursor.execute(dataquery, (first_date, last_date))

        alldata = numpy.empty((0,2))

        for (Date, Close, Intrest) in cursor:
            #print("On {:%d %b %Y} the data was: {}".format(Date, Close))
            alldata = numpy.append(alldata, [Date,float(Close),float(Intrest)])
        
        marketquery = ("SELECT market_id FROM Indexes WHERE index_name = '{}'".format(table))
        cursor.execute(marketquery)
        for id in cursor:
            market_id = id[0]
        print(market_id)
        cursor.close
        indexquery = ("SELECT index_name FROM Indexes where market_id = {}".format(market_id))
        cursor.execute(indexquery)
        print("These indexes are in the same market:")
        for index in cursor:
            print(index[0])

        cursor.close()
        cnx.close()
        print(alldata, numpy.shape(alldata))
        alldata = numpy.reshape(alldata, [int(len(alldata)/3), 3])
        print(alldata, numpy.shape(alldata))
        return alldata

    def getEpisodeLength(self):
        return len(self.trainingdata) - self.windowSize - 1

    def reset(self, specialParameter = 0.0, hardReset = False):
        #TODO store results
        
        self.ws.cell(row=self.episode + 1, column= self.specialparameterIterator + 2, value= self.profit)
        self.ws.cell(row=self.episode + 1, column=1, value=self.episode)
        self.ws.cell(row=self.episode + 1, column=3, value=self.epsilon)
        self.ws.cell(row=self.episode + 1, column=4, value=self.startDate)
        if hardReset:
            self.specialparameterIterator += 1
            self.specialparameter = specialParameter
            self.episode = 1
            self.ws.cell(row=1, column= self.specialparameterIterator + 2, value=specialParameter)
        
        self.wb.save(self.resultsfile)
        self.episode += 1
        self.time = 0
        self.profit = 0.0
        self.invested = 0.0
        self.inventory.clear()

    def getState(self):
        """set instrument value and format/return current state to the agent"""
        _, self.instrumentValue,_ = self.trainingdata[self.time+self.windowSize]
        data = self.trainingdata[self.time:self.time+self.windowSize]
        data = numpy.swapaxes(data, 0, 1)[1:]
        data = numpy.array(list(data), dtype="float32")
        #data = numpy.concatenate(data)
        sellprofit = data[1]-self.inventory[0] if len(self.inventory) else numpy.zeros(self.windowSize, dtype="float32")
        #data = numpy.append(data, [profitsell, self.invested, float(len(self.inventory))]).astype('float32')
        data = numpy.vstack([data, sellprofit])

        return data

    def action(self, action):
        reward = 0.0
        if action == 0:
            self.inventory.append(self.instrumentValue)
            self.invested += self.instrumentValue
        elif action == 1 and self.inventory:
            self.invested -= self.inventory[0]
            reward = self.instrumentValue - self.inventory[0]
            self.inventory = self.inventory[1:]


        self.profit += reward
        print("On t={} the profit is: {:.2f}".format(self.time, self.profit))
        self.time += 1
        done = False
        if self.time  == self.getEpisodeLength() - 1:
            done = True
        next_state = self.getState()
        
        return next_state, reward, done
