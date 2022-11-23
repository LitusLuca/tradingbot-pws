import math
from datetime import datetime, timedelta

import mysql.connector
import numpy


from _env import password, user
from openpyxl import Workbook, load_workbook
from os.path import exists


class StockSimulation:
    def __init__(self, instrument, timeSpan, episodeStart, results) -> None:
        self.time = 0
        self.windowSize = 32
        self.inventory = []
        self.invested = 0.0
        self.profit = 0.0
        self.instrument = instrument #TODO
        self.actionSpace = 3
        self.inputSpace = self.windowSize + 2
        self.trainingdata = self._getData(instrument, episodeStart-timedelta(days=self.windowSize), timeSpan)
        _, self.instrumentValue = self.trainingdata[self.time]
        print(self.instrumentValue)

        if exists(results):
            self.wb = load_workbook(results)
        else:
            self.wb = Workbook()
        self.ws = self.wb.active
        self.specialparameter = 0.0
        self.specialparameterIterator = 0
        self.episode = 1
        self.ws["A1"] = "episode \\ parameter"
        self.ws["B1"] = self.specialparameter
        self.wb.save(results)
        self.resultsfile = results


    def _getData(self, instrument, startDate, length):
        cnx = mysql.connector.connect(user=user,password=password, database='indexes')
        cursor = cnx.cursor()

        table = instrument

        dataquery = ("SELECT Date, Close FROM `{}`"
            "WHERE Date BETWEEN %s AND %s".format(table))

        first_date = startDate
        last_date = first_date + timedelta(days=length)

        cursor.execute(dataquery, (first_date, last_date))

        alldata = []

        for (Date, Close) in cursor:
            print("On {:%d %b %Y} the data was: {}".format(
            Date, Close))
            alldata.append((Date,float(Close)))
        
        marketquery = ("SELECT market_id FROM Indexes WHERE index_name = '{}'".format(table))
        cursor.execute(marketquery)
        for id in cursor:
            market_id = id[0]
        print(market_id)
        cursor.close
        indexquery = ("SELECT index_name FROM Indexes where market_id = {}".format(market_id))
        cursor.execute(indexquery)
        print("These indexes are in the same market:")
        for index in cursor:
            print(index[0])

        cursor.close()
        cnx.close()
        return alldata

    def getEpisodeLength(self):
        return len(self.trainingdata) - self.windowSize

    def reset(self, specialParameter = 0.0, hardReset = False):
        #TODO store results
        
        self.ws.cell(row=self.episode + 1, column= self.specialparameterIterator + 2, value= self.profit)
        self.ws.cell(row=self.episode + 1, column=1, value=self.episode)
        if hardReset:
            self.specialparameterIterator += 1
            self.specialparameter = specialParameter
            self.episode = 1
            self.ws.cell(row=1, column= self.specialparameterIterator + 2, value=specialParameter)
        
        self.wb.save(self.resultsfile)
        self.episode += 1
        self.time = 0
        self.profit = 0.0
        self.invested = 0.0
        self.inventory.clear()

    def getState(self):
        """set instrument value and format/return current state to the agent"""
        _, self.instrumentValue = self.trainingdata[self.time+self.windowSize]
        data = self.trainingdata[self.time:self.time+self.windowSize]
        data = list(list(zip(*data))[1])
        data.extend([self.profit, self.invested])
        return numpy.reshape(data, [1, self.inputSpace])

    def action(self, action):
        reward = 0.0
        if action == 0:
            self.inventory.append(self.instrumentValue)
            self.invested += self.instrumentValue
        elif action == 1 and self.inventory:
            self.invested -= self.inventory[0]
            reward = self.instrumentValue - self.inventory[0]
            self.inventory = self.inventory[1:]


        self.profit += reward
        print("On t={} the profit is: {:.2f}".format(self.time, self.profit))
        self.time += 1
        if self.time  == self.getEpisodeLength():
            return [], reward, True
        next_state = self.getState()
        done = False
        return next_state, reward, done
